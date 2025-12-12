# excel.py — create excel exports and apply basic styling with openpyxl
import os
from openpyxl import load_workbook
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import current_app

def style_excel_with_title(path: str, title: str):
    wb = load_workbook(path)
    ws = wb.active
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="14365A")
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    zebra_fill = PatternFill("solid", fgColor="F6F6F6")
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 1 or max_col < 1:
        wb.save(path); return
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    tcell = ws.cell(row=1, column=1); tcell.value = title; tcell.font = title_font; tcell.alignment = center
    for c in range(1, max_col+1):
        cell = ws.cell(row=2, column=c); cell.font = header_font; cell.fill = header_fill; cell.alignment = center; cell.border = border
    for r in range(3, max_row+2):
        is_zebra = ((r-2)%2==0)
        for c in range(1, max_col+1):
            cell = ws.cell(row=r, column=c); cell.alignment = center; cell.border = border
            if is_zebra: cell.fill = zebra_fill
    ws.freeze_panes = ws["A3"]
    for c in range(1, max_col+1):
        col_letter = get_column_letter(c)
        max_len = 0
        for r in range(1, max_row+2):
            v = ws.cell(row=r, column=c).value
            txt = str(v) if v is not None else ""
            if len(txt) > max_len: max_len = len(txt)
        ws.column_dimensions[col_letter].width = max(10, max_len + 2)
    wb.save(path)

def save_df_to_excel(df, excel_path, title="Report"):
    df.to_excel(excel_path, index=False)
    style_excel_with_title(excel_path, title)
